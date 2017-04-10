#! /usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import shutil
import datetime
import win32api # dialog box

from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl import load_workbook

source = u'T:\Projekty\Indukcyjności\Zxx SZABLON'
sciezka = u'T:\Projekty\Indukcyjności'



# -------------------------------------- AKTUALIZACJA KATALOGU
def akt_kat(directory):
	os.chdir(os.path.dirname(directory))
	temp = directory.split("\\")
	temp.pop()

	temp_kli = temp[len(temp)-2]
	temp_proj = temp[len(temp)-1]
	temp2 = temp[len(temp)-2]
	nr_proj = temp_proj.split(" ")[0]
	nazwa = " ".join(temp_proj.split(" ")[1:])
	data = datetime.date.today()
	klient = " ".join(temp_kli.split(" ")[1:])

	# aktualizacja nazw plików nowostworzonego katalogu
	try:
		os.rename('PWxxx_xx czytaj nowy.txt', '%s czytaj nowy.txt' %nr_proj)
		os.rename('PWxxx_xx_IND_(Klient_data_nazwa).xlsx', '%s_%s_%s_%s.xlsx' %(nr_proj, klient, str(data), nazwa))
		with open('%s czytaj nowy.txt' %nr_proj, "a") as myfile:
			myfile.write("JG: %s" %data)
	except:
		win32api.MessageBox(0, "Coś poszlo nie tak", "Aktualzacja zawartości katalogu", 0x00001000)# pops on TOP
		raise
		
	# --------------------------------------- aktualizacja wyceny
	try:
		temp.append(('%s_%s_%s_%s.xlsx' %(nr_proj, klient, str(data), nazwa)))
		dir_wyc = "/".join(temp)

		wb = load_workbook(dir_wyc) 
		ws = wb.active 
		ws['L2'] = nazwa
		ws['E2'] = nr_proj
		ws['E3'] = '1'
		ws['E4'] = data
		ws['E5'] = 'Jakub Grzegulski'
		wb.save(dir_wyc)
		temp.pop()
	except:
		win32api.MessageBox(0, "Coś poszlo nie tak", "Aktualzacja wyceny", 0x00001000)# pops on TOP
		raise
	# --------------------------------------- aktualizacja naklejki
	try:
		temp.append(('naklejka.xlsx'))
		dir_nakl = "/".join(temp)
		wb = load_workbook(dir_nakl) 
		ws = wb.active 
		ws['A3'] = nazwa
		wb.save(dir_nakl)
		temp.pop()
	except:
		win32api.MessageBox(0, "Coś poszlo nie tak", "Aktualzacja naklejki", 0x00001000)# pops on TOP
		raise

	# --------------------------------------- aktualizacja KWP-03
	try:
		temp.append(u'prototyp')
		temp.append(u'KWP-03.xlsx')
		dir_proj = "/".join(temp)
		temp.pop()
		temp.append(u'logo.jpg')
		dir_im = "/".join(temp)

		wb = load_workbook(dir_proj) 
		ws = wb.active 
		ws['K1'] = datetime.date.today()
		ws['D4'] = klient
		ws['D5'] = nr_proj
		ws['D6'] = nazwa

		img = Image(dir_im)
		ws.add_image(img)

		wb.save(dir_proj)

		os.chdir(os.path.dirname(dir_proj))
		os.rename('KWP-03.xlsx', 'KWP-03 %s.xlsx' %nazwa)
	except:
		win32api.MessageBox(0, "Coś poszlo nie tak", "Aktualzacja KWP-03", 0x00001000)# pops on TOP
		raise
