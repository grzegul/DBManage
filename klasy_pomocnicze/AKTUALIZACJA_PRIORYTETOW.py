#! /usr/bin/env python
# -*- coding: utf-8 -*-

import re
import os
import sys
import imp
import shutil
import datetime
import win32api # dialog box

from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import numbers, is_date_format
from openpyxl.styles.styleable import StyleableObject



source = u'T:\Projekty\Indukcyjności\Zxx SZABLON'
sciezka = u'T:\Projekty\Indukcyjności\\' + "\\".join((os.path.dirname(sys.argv[0])).split("\\")[3:])
#os.chdir(os.path.dirname(sys.argv[0]))	# wykonanie w katalogu, w którym się znajduje



# -------------------------------------- AKTUALIZACJA PRIORYTETÓW
#os.chdir(os.path.dirname(u'J:/drop/trafa/OBLICZENIA')) #zmiana ścieżki
dir_PRI = u'J:/drop/trafa/OBLICZENIA/PRIORYTETY MAGNETYKI.xlsx'
wb = load_workbook(dir_PRI, data_only=False) 
ws = wb.active
	
num = ws.max_row + 1
ws['A%d' %(num)] = num - 1
ws['B%s' %(num)] = data
ws['D%s' %(num)] = 'w trakcie'
ws['E%s' %(num)] = nr_proj
ws['F%s' %(num)] = klient
ws['G%s' %(num)] = nazwa
ws['H%s' %(num)] = 'wycenaa
ws['I%s' %(num)] = 'JG'

wb.save(dir_PRI)

